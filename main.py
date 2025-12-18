#!/usr/bin/env python3
"""
Email Validation Tool
This script performs email validation using DNS checks (NO SMTP).
"""

import csv
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill
from verifier import EmailVerifier


INPUT_FILE = 'emails_input.txt'
OUTPUT_TXT = 'emails_output.txt'
OUTPUT_CSV = 'emails_output.csv'
OUTPUT_XLSX = 'emails_output.xlsx'


def main():
    """Main entry point for email validation."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(base_dir, INPUT_FILE)
    
    if not os.path.exists(input_path):
        print(f"Error: {INPUT_FILE} not found.")
        sys.exit(1)
    
    with open(input_path, 'r', encoding='utf-8') as f:
        emails = [line.strip() for line in f if line.strip()]
    
    if not emails:
        print("No emails to validate.")
        sys.exit(0)
    
    verifier = EmailVerifier()
    
    print(f"Starting validation for {len(emails)} emails...")
    print("=" * 60)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Validation"
    ws.append(['Email', 'Status'])
    
    # Define cell styles
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Output paths
    csv_path = os.path.join(base_dir, OUTPUT_CSV)
    txt_path = os.path.join(base_dir, OUTPUT_TXT)
    xlsx_path = os.path.join(base_dir, OUTPUT_XLSX)
    
    # Process emails and write results
    with open(csv_path, 'w', newline='', encoding='utf-8') as f_csv, \
         open(txt_path, 'w', encoding='utf-8') as f_txt:
        
        writer = csv.writer(f_csv)
        writer.writerow(['Email', 'Status'])
        
        valid_count = 0
        invalid_count = 0
        risky_count = 0
        
        for i, email in enumerate(emails, 1):
            print(f"[{i}/{len(emails)}] Validating: {email}")
            status = verifier.verify(email)
            
            # Count statuses
            if status == 'VALID':
                valid_count += 1
            elif status == 'INVALID':
                invalid_count += 1
            elif status == 'RISKY':
                risky_count += 1
            
            # Write to CSV
            writer.writerow([email, status])
            
            # Write to TXT
            f_txt.write(f"{email}: {status}\n")
            
            # Write to Excel
            ws.append([email, status])
            
            # Apply red fill for INVALID or RISKY
            if status in ('INVALID', 'RISKY'):
                for col in range(1, 3):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.fill = red_fill
            
            print(f"    → Status: {status}")
    
    # Save Excel file
    wb.save(xlsx_path)
    
    print("=" * 60)
    print("Validation complete.")
    print(f"  VALID:   {valid_count}")
    print(f"  INVALID: {invalid_count}")
    print(f"  RISKY:   {risky_count}")
    print(f"\nOutputs:")
    print(f"  - {OUTPUT_TXT}")
    print(f"  - {OUTPUT_CSV}")
    print(f"  - {OUTPUT_XLSX}")
    print("=" * 60)


if __name__ == "__main__":
    main()